import os 
import openpyxl
from openpyxl import load_workbook
from string import ascii_uppercase

os.system('cls')

dateiName = "test.xlsx"

def neueExcelTabelleErstellen(_dateiName):
    #Leeren Workbook erstellen
    wb = openpyxl.Workbook()

    #Sheet für Übungen und Fortschritt erstellen
    übungsSheet = wb.active
    übungsSheet.title = 'Uebungen'
    datenSheet = wb.create_sheet('Trainingsdaten')

    #Beim erstmaligen Erstellen des Excels noch nur ges. Wiederholungen. Beim Verarbeiten durch Athena müssen dann noch die einzelnen Wiederholungen der Sätze dazukommen.
    headerÜbungsSheet = ['Uebung', 'Saetze', 'Wiederholungen', 'Beschreibung']
    headerDatenSheet = ['Datum', 'Uebung', 'Saetze', 'Gesamte Wiederholungen']
    

    # Überschriften in die oberste Zeile der Exceltabelle eintragen
    for i in range(0, len(headerÜbungsSheet)):
        übungsSheet[str(ascii_uppercase[i]) + str(1)] = headerÜbungsSheet[i]
    
    for i in range(0, len(headerDatenSheet)):
        datenSheet[str(ascii_uppercase[i]) + str(1)] = headerDatenSheet[i]

    #Workbook abspeichern
    wb.save(filename=_dateiName)


def übungHinzufügen(_dateiName, _übung, _sets, _reps, _beschreibung):
    #excel übungsSheet öffnen
    wb = load_workbook(filename = _dateiName)
    sheet = wb.active
    #checken welches column noch nicht eingetragen wurde

    #auf das nächste freie column die attribute eintragen
    liste = [_übung, _sets, _reps, _beschreibung]
    sheet.append(liste)
    print(sheet.max_row)
    #excel speichern
    wb.save(filename= _dateiName)
    

neueExcelTabelleErstellen(dateiName)
übungHinzufügen(dateiName, 'Erste Uebung', 3, 8, 'Hier steht eine Beschreibung der Uebung')
übungHinzufügen(dateiName, 'Zweite Uebung', 3, 10, 'Hier steht die zweite Beschreibung') 